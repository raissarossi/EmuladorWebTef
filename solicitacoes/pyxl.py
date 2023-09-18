import time
import os
import pandas as pd
from openpyxl import Workbook
import openpyxl
import datetime
import json
import shutil

# EXEMPLO:
with open("configs.txt", "r", encoding="utf-8") as read:
        configs = json.load(read)
DATABASE = configs['DATABASE']
PASTA_MODELO = configs['PASTA_MODELO']
PASTA_DESTINO = configs['PASTA_DESTINO']

def copiar_pasta_modelo(id):

    # entra para pegar a pasta modelo 
    nomes = os.listdir(PASTA_MODELO)

    pastaAnoAtual = "9999-23"

    # pega a pasta modelo para copiar e o id do ano atual
    for i in nomes:
        if i.__contains__("MODELO") and i.__contains__("-"):
            pastaAnoAtual = i
    # junta o id com ano
    idAno = str(id)+pastaAnoAtual[pastaAnoAtual.find("-"):pastaAnoAtual.find(" ")]

    # define a pasta de destino ja com o ano e id
    destino = PASTA_DESTINO+"\\"+idAno

    # copia a pasta modelo do modelo para a pasta das solicitações !! A pasta destino não deve existir, ela sera criada no momento em que for copiado !!
    shutil.copytree(PASTA_MODELO+"\\"+pastaAnoAtual, destino)
    

    novo_caminho_arquivo = os.path.join(destino, idAno+" Custo por Conjunto"+".xlsm")

    # renomeia o xlsm do excel
    os.rename(destino+"\\9999-23 Custo por Conjunto.xlsm", novo_caminho_arquivo)


# TESTE MANUAL
# def send():
#     wb_obj = openpyxl.load_workbook(filename=DATABASE)
#     sheet_obj = wb_obj.active
#     while True:
#         try:
#             wb_obj.save(filename=DATABASE)
#             break
#         except PermissionError:
#             print('trying...')
#             time.sleep(5)
#     while True:
#         try:
#             wb_obj = openpyxl.load_workbook(filename=DATABASE)
#             break
#         except EOFError:
#             print('teste')
#             time.sleep(5)

#     sheet_obj = wb_obj.active

#     lastId = get_last_id(sheet_obj)

#     numero_da_linha = sheet_obj.max_row + 1  # Obtém o número da última linha + 1

#     coluna_desejada = 'A'  # Coluna desejada para determinar a última linha preenchida
#     ultima_linha = None

#     for cell in sheet_obj[coluna_desejada]:
#         if cell.value is not None:
#             ultima_linha = cell.row

#     if ultima_linha is not None:
#         numero_da_linha = ultima_linha + 1
#     else:
#         numero_da_linha = 1  # Se não houver nenhuma linha preenchida, começar na linha 1

#     valores = [int(lastId)+1, 'EU4K1', 'ESSEMSM', 5555, '10/12/2023', 'Aquele la', 666666, 'Varios Consumo', 'Precário', 'Materia prima do fornecedor', '20/12/2023', 'BEM PRECISA', 'P4NACO7A', 'NA', 'Não', 'NA']

#     for coluna, valor in enumerate(valores, start=1):
#         sheet_obj.cell(row=numero_da_linha, column=coluna).value = valor
    
#     wb_obj.save(filename=DATABASE)
#     # print(sheet_obj.__dict__)
    
def send_with_django(dados):

    # tenta abrir o banco em excel e salvar para verificar se o arquivo ja esta aberto
    try:
        wb_obj = openpyxl.load_workbook(filename=DATABASE)
    except Exception as e:
        return -1
    sheet_obj = wb_obj.active
    while True:
        try:
            # tentar salvar para verificar se o arquivo ja esta aberto
            wb_obj.save(filename=DATABASE)
            break
        except PermissionError:
            # se estiver aberto ele espera e tenta novamente
            time.sleep(5)

    # a partir do momento em que é fechado ele inicia e começa a colocar os dados
    while True:
        try:
            wb_obj = openpyxl.load_workbook(filename=DATABASE)
            break
        except EOFError:
            
            time.sleep(5)

    sheet_obj = wb_obj.active

    lastId = get_last_id(sheet_obj)

    numero_da_linha = sheet_obj.max_row + 1  # Obtém o número da última linha + 1

    coluna_desejada = 'A'  # Coluna desejada para determinar a última linha preenchida
    ultima_linha = None

    for cell in sheet_obj[coluna_desejada]:
        if cell.value is not None:
            ultima_linha = cell.row

    if ultima_linha is not None:
        numero_da_linha = ultima_linha + 1
    else:
        numero_da_linha = 1  # Se não houver nenhuma linha preenchida, começar na linha 1

    # valores que serao inseridos
    valores =   [ 
                int(lastId)+1,
                dados["user"],
                dados["setor"],
                dados["ramal"],
                dados["data"],
                dados["local_debito_tipo"],
                dados["local_debito_numero"],
                dados["tipo_debito"],
                dados["categoria"],
                dados["tipo_de_servico"],
                dados["prazo_pretendido"],
                dados["descricao"],
                dados["n_desenho"],
                dados["grupo"],
                dados["solicitacao_semelhante"],
                dados["n_solicitacao_executada"],
                ]
    # percorre a lista enquanto coloca as informações no banco de dados 
    for coluna, valor in enumerate(valores, start=1):
        sheet_obj.cell(row=numero_da_linha, column=coluna).value = valor
    # print("escreveu")
    try :
        # print("salvou")

        # se conseguir salvar ele continua se não ele para
        wb_obj.save(filename=DATABASE)
    except:
        return -2
    try:
        # copia a pasta modelo
        copiar_pasta_modelo(int(lastId)+1)
        # print("copiou")
    except:
        return -1
    # se ele não der erro ele volta o id atual
    return int(lastId)+1
    
    

def get_last_id(sheet_obj):

    m_row = sheet_obj.max_row
    
    lastId = 1
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        if (cell_obj.value == None):
            continue
        lastId = cell_obj.value
    return lastId


# if __name__ == "__main__":
#     wb_obj = openpyxl.load_workbook(filename="951-23 Custo por Conjunto.xlsm")
#     wa = wb_obj.active
#     wa["B1"] = "123-23"
#     wb_obj.save("951-23 Custo por Conjunto.xlsm")